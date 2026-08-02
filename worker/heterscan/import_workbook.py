from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from .normalize import normalized_key
from .supabase import SupabaseRepository


def _rows(path: Path, sheet_name: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [str(cell.value) for cell in sheet[4]]
    for values in sheet.iter_rows(min_row=5, values_only=True):
        if not any(value is not None for value in values):
            continue
        yield dict(zip(headers, values, strict=False))


def import_workbook(path: Path, repository: SupabaseRepository) -> tuple[int, int]:
    official_rows = [
        {
            "city_id": str(row["locality_code"]), "street_code": str(row["street_code"]),
            "official_code": str(row.get("official_code") or "") or None,
            "official_name": str(row["official_street_name"]).strip(),
        }
        for row in _rows(path, "Official_Streets")
    ]
    official_count = repository.upsert_many(
        "official_streets", official_rows, conflict="city_id,street_code", batch_size=400
    )
    identifiers = repository._rest("GET", "official_streets?select=id,city_id,street_code").json()
    id_by_key = {(str(row["city_id"]), str(row["street_code"])): row["id"] for row in identifiers}
    aliases = []
    for row in _rows(path, "Search_Variants"):
        key = (str(row["locality_code"]), str(row["street_code"]))
        official_id = id_by_key.get(key)
        alias = str(row.get("search_variant") or "").strip()
        if not official_id or not alias:
            continue
        aliases.append(
            {
                "official_street_id": official_id, "alias": alias, "alias_normalized": normalized_key(alias),
                "priority": 0 if row.get("variant_type") == "official" else 100,
            }
        )
    alias_count = repository.upsert_many(
        "street_aliases", aliases, conflict="official_street_id,alias_normalized", batch_size=400
    )
    return official_count, alias_count


def main() -> None:
    project_root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Import the verified five-authority street workbook")
    parser.add_argument(
        "workbook", nargs="?", type=Path,
        default=project_root / "data" / "5 ATHORITIES_STREETS_DATA_FOR_PILOT_TEST.xlsx",
    )
    args = parser.parse_args()
    repository = SupabaseRepository()
    try:
        official, aliases = import_workbook(args.workbook.resolve(), repository)
        print(f"Imported {official:,} official streets and {aliases:,} search variants")
    finally:
        repository.close()


if __name__ == "__main__":
    main()
