from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="123B70")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _display(value: Any) -> Any:
    if value is None or value == "":
        return "לא ידוע"
    if isinstance(value, (dict, list)):
        return str(value)
    return value


def _sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], headers: list[tuple[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.rightToLeft = True
    sheet.append([label for _, label in headers])
    for row in rows:
        sheet.append([_display(row.get(key)) for key, _ in headers])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True, readingOrder=2)
        source_cell = row[-1] if row else None
        if source_cell and isinstance(source_cell.value, str) and source_cell.value.startswith("http"):
            source_cell.hyperlink = source_cell.value
            source_cell.style = "Hyperlink"
    for index, (_, label) in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(55, max(14, len(label) + 5))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_report(run: dict[str, Any], results: list[dict[str, Any]], units: list[dict[str, Any]]) -> tuple[bytes, str]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = [{
        "city_name": run["city_name"], "date_from": run["date_from"], "date_to": run["date_to"],
        "status": run["status"], "applications_found": len(results),
        "permits_found": sum(bool(row.get("is_permit_issued")) for row in results),
        "units_total": len(units), "units_completed": sum(row["status"] == "completed" for row in units),
    }]
    _sheet(workbook, "סיכום", summary, [
        ("city_name", "עיר"), ("date_from", "מתאריך"), ("date_to", "עד תאריך"), ("status", "סטטוס"),
        ("applications_found", "בקשות שנמצאו"), ("permits_found", "היתרים שנמצאו"),
        ("units_total", "יחידות חיפוש"), ("units_completed", "יחידות שהושלמו"),
    ])
    common_headers = [
        ("address", "כתובת"), ("application_number", "מספר בקשה"), ("building_file_number", "מספר תיק בניין"),
        ("block_number", "גוש"), ("parcel_number", "חלקה"), ("application_type", "סוג בקשה"),
        ("work_description", "תיאור עבודה"), ("submission_date", "תאריך הגשה"),
        ("approval_date", "תאריך אישור"), ("permit_number", "מספר היתר"),
        ("permit_issue_date", "תאריך הפקת היתר"), ("permit_status_original", "סטטוס מקורי"),
        ("permit_confidence", "רמת אמינות"), ("source_url", "קישור מקור"),
    ]
    permits = [row for row in results if row.get("is_permit_issued")]
    approvals = [row for row in results if row.get("is_approved")]
    _sheet(workbook, "היתרים שנמצאו", permits, common_headers)
    _sheet(workbook, "בקשות שאושרו", approvals, common_headers)
    _sheet(workbook, "כל התוצאות", results, common_headers)
    _sheet(workbook, "יחידות חיפוש", units, [
        ("sequence", "מספר"), ("unit_key", "יחידה"), ("status", "סטטוס"), ("attempts", "ניסיונות"),
        ("result_count", "תוצאות"), ("completed_at", "זמן סיום"), ("error_message", "שגיאה"),
    ])
    errors = [row for row in units if row["status"] in ("failed", "requires_review")]
    _sheet(workbook, "שגיאות", errors, [
        ("sequence", "מספר"), ("unit_key", "יחידה"), ("status", "סטטוס"),
        ("attempts", "ניסיונות"), ("error_message", "פירוט"),
    ])
    stream = BytesIO()
    workbook.save(stream)
    payload = stream.getvalue()
    return payload, hashlib.sha256(payload).hexdigest()
