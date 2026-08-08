from io import BytesIO

from openpyxl import load_workbook

from heterscan.reporting import build_report


def test_report_includes_pending_applications_with_source_link() -> None:
    results = [
        {
            "application_number": "20250001",
            "permit_number": "2025-100",
            "permit_issue_date": "2025-12-15",
            "permit_status_original": "היתר הופק",
            "is_approved": True,
            "is_permit_issued": True,
            "source_url": "https://example.test/issued",
        },
        {
            "application_number": "20250002",
            "is_approved": False,
            "is_permit_issued": False,
            "source_url": "https://example.test/pending",
        },
    ]
    run = {
        "city_name": "פתח תקווה",
        "date_from": "2025-12-01",
        "date_to": "2025-12-31",
        "status": "completed",
    }

    payload, _ = build_report(run, results, [])
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook["בקשות והיתרים"]
    headers = {cell.value: cell.column for cell in sheet[1]}

    assert sheet.max_row == 3
    assert sheet.cell(3, headers["סטטוס"]).value == "טרם אושר"
    source_cell = sheet.cell(3, headers["קישור מקור"])
    assert source_cell.value == "https://example.test/pending"
    assert source_cell.hyperlink.target == "https://example.test/pending"
    assert workbook["היתרים שנמצאו"].max_row == 2
